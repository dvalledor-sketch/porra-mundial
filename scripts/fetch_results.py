"""
fetch_results.py — descarga resultados del Mundial 2026 desde football-data.org
y actualiza data/resultados_reales.xlsx automáticamente.

Requiere la variable de entorno FOOTBALL_DATA_TOKEN con tu token de
https://www.football-data.org/client/register (el plan gratuito incluye
el Mundial; límite 10 peticiones/minuto, de sobra).

Solo llama a la API cuando un partido debería haber terminado y aún no
tiene resultado (ver _toca_llamar). FORCE_FETCH=1 o --force fuerzan la
llamada (p. ej. en ejecuciones manuales del workflow).

Uso manual:
    FOOTBALL_DATA_TOKEN=xxx python scripts/fetch_results.py [--force]
"""

from __future__ import annotations

import json
import os
import sys
from pathlib import Path

import requests
from openpyxl import load_workbook

# ── rutas ─────────────────────────────────────────────────────────────────────
RAIZ        = Path(__file__).resolve().parent.parent
MATCHES_F   = RAIZ / "data" / "matches.json"
REALES_F    = RAIZ / "data" / "resultados_reales.xlsx"
REALES_ELIM = RAIZ / "data" / "resultados_reales_elim.xlsx"
BRACKET_F   = RAIZ / "data" / "bracket.json"
HORARIOS_F  = RAIZ / "data" / "horarios.json"

# ── API (football-data.org v4) ────────────────────────────────────────────────
API_BASE    = "https://api.football-data.org/v4"
API_TOKEN   = os.environ.get("FOOTBALL_DATA_TOKEN", "")
COMPETITION = "WC"   # FIFA World Cup

# ── mapeo nombre API (inglés) → nombre interno (español) ──────────────────────
NOMBRES: dict[str, str] = {
    "Mexico": "México", "South Africa": "Sudáfrica",
    "South Korea": "Corea del Sur", "Korea Republic": "Corea del Sur",
    "Czechia": "Chequia", "Czech Republic": "Chequia",
    "Canada": "Canadá", "Switzerland": "Suiza",
    "Qatar": "Catar", "Bosnia": "Bosnia y Herzegovina",
    "Bosnia and Herzegovina": "Bosnia y Herzegovina",
    "Brazil": "Brasil", "Morocco": "Marruecos",
    "Scotland": "Escocia", "Haiti": "Haití",
    "USA": "Estados Unidos", "United States": "Estados Unidos",
    "Paraguay": "Paraguay", "Australia": "Australia",
    "Turkey": "Turquía", "Türkiye": "Turquía",
    "Germany": "Alemania", "Curacao": "Curazao", "Curaçao": "Curazao",
    "Ivory Coast": "Costa de Marfil", "Côte d'Ivoire": "Costa de Marfil",
    "Cote d'Ivoire": "Costa de Marfil", "Ecuador": "Ecuador",
    "Netherlands": "Países Bajos", "Japan": "Japón",
    "Tunisia": "Túnez", "Sweden": "Suecia",
    "Belgium": "Bélgica", "Egypt": "Egipto",
    "Iran": "Irán", "IR Iran": "Irán", "New Zealand": "Nueva Zelanda",
    "Spain": "España", "Cape Verde": "Cabo Verde", "Cabo Verde": "Cabo Verde",
    "Saudi Arabia": "Arabia Saudí", "Uruguay": "Uruguay",
    "France": "Francia", "Senegal": "Senegal",
    "Norway": "Noruega", "Iraq": "Irak",
    "Argentina": "Argentina", "Algeria": "Argelia",
    "Austria": "Austria", "Jordan": "Jordania",
    "Portugal": "Portugal", "Colombia": "Colombia",
    "Uzbekistan": "Uzbekistán", "DR Congo": "RD Congo", "Congo DR": "RD Congo",
    "England": "Inglaterra", "Croatia": "Croacia",
    "Ghana": "Ghana", "Panama": "Panamá", "Panamá": "Panamá",
}


def _es(api_name: str) -> str:
    """Convierte nombre API a nombre español."""
    return NOMBRES.get(api_name, api_name)


def _equipo_es(team: dict) -> str:
    """Nombre español a partir del objeto team de la API (name o shortName)."""
    for key in ("name", "shortName"):
        n = team.get(key)
        if n and n in NOMBRES:
            return NOMBRES[n]
    return team.get("name") or team.get("shortName") or "?"


def _api_get(endpoint: str, params: dict | None = None) -> dict:
    r = requests.get(
        f"{API_BASE}/{endpoint}",
        headers={"X-Auth-Token": API_TOKEN},
        params=params or {},
        timeout=20,
    )
    if r.status_code != 200:
        # football-data devuelve {"message": "...", "errorCode": ...} en errores
        try:
            detalle = r.json().get("message", r.text[:200])
        except Exception:
            detalle = r.text[:200]
        raise RuntimeError(f"football-data.org HTTP {r.status_code}: {detalle}")
    return r.json()


def _get_all_matches() -> list[dict]:
    """Devuelve TODOS los partidos del Mundial 2026 en una sola llamada."""
    data = _api_get(f"competitions/{COMPETITION}/matches")
    return data.get("matches", [])


TERMINADO = {"FINISHED", "AWARDED"}


def _build_match_index() -> dict[tuple[str, str], str]:
    """Construye dict (local_es, visitante_es) → match_id a partir de matches.json."""
    with open(MATCHES_F, encoding="utf-8") as f:
        data = json.load(f)
    return {(m["local"], m["visitante"]): m["id"] for m in data["partidos"]}


def _update_excel(archivo: Path, resultados: dict[str, tuple[int, int]]) -> int:
    """
    Actualiza el Excel de resultados reales con los marcadores obtenidos.
    Solo rellena celdas que estén vacías (no sobreescribe correcciones manuales).
    Devuelve el número de celdas actualizadas.
    """
    if not archivo.exists():
        print(f"⚠️  No encontrado: {archivo}")
        return 0

    wb = load_workbook(archivo)
    actualizados = 0

    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        for row in ws.iter_rows(min_row=1):
            cells = [c.value for c in row]
            padded = cells + [None] * (10 - len(cells))
            match_id = padded[2]
            if not isinstance(match_id, str):
                continue
            match_id = match_id.strip()
            if match_id not in resultados:
                continue
            gl, gv = resultados[match_id]
            col_d = row[3] if len(row) > 3 else None
            col_e = row[4] if len(row) > 4 else None
            if col_d is not None and col_d.value is None:
                col_d.value = gl
                actualizados += 1
            if col_e is not None and col_e.value is None:
                col_e.value = gv
                actualizados += 1

    if actualizados:
        wb.save(archivo)
    return actualizados


def _resultados_existentes() -> set[str]:
    """IDs de partido que ya tienen resultado en los Excel de resultados reales."""
    ids: set[str] = set()
    for archivo in (REALES_F, REALES_ELIM):
        if not archivo.exists():
            continue
        wb = load_workbook(archivo, read_only=True)
        for ws in wb:
            for row in ws.iter_rows():
                cells = [c.value for c in row]
                padded = cells + [None] * (10 - len(cells))
                mid, gl, gv = padded[2], padded[3], padded[4]
                if isinstance(mid, str) and gl is not None and gv is not None:
                    ids.add(mid.strip())
    return ids


def _toca_llamar() -> tuple[bool, str]:
    """Decide si hay que llamar a la API: solo cuando un partido debería haber
    terminado (kickoff + 105 min) y aún no tiene resultado. Así la porra se
    actualiza al final de cada partido sin abusar de la API."""
    import datetime as dt
    now = dt.datetime.now(dt.timezone.utc)

    if os.environ.get("FORCE_FETCH") == "1" or "--force" in sys.argv:
        return True, "forzado (FORCE_FETCH / --force)"

    con_resultado = _resultados_existentes()

    # Fase de grupos: horarios oficiales (data/horarios.json, UTC)
    if HORARIOS_F.exists():
        horarios = json.loads(HORARIOS_F.read_text(encoding="utf-8"))
        for mid, iso in horarios.items():
            if mid in con_resultado:
                continue
            k = dt.datetime.fromisoformat(iso)
            if k + dt.timedelta(minutes=105) <= now <= k + dt.timedelta(hours=6):
                return True, f"{mid} debería haber terminado (kickoff {iso})"

    # Eliminatorias: fechas reales del bracket descargado de la API
    if BRACKET_F.exists():
        try:
            rounds = json.loads(BRACKET_F.read_text(encoding="utf-8")).get("rounds", {})
        except Exception:
            rounds = {}
        for fixes in rounds.values():
            for f in fixes:
                if f.get("status") in ("FT", "AET", "PEN") or not f.get("date"):
                    continue
                try:
                    k = dt.datetime.fromisoformat(f["date"].replace("Z", "+00:00"))
                except ValueError:
                    continue
                if k + dt.timedelta(minutes=105) <= now <= k + dt.timedelta(hours=7):
                    return True, f"{f.get('team1')}–{f.get('team2')} debería haber terminado"

    # Red de seguridad: una pasada completa al día (06:00–06:14 UTC)
    if now.hour == 6 and now.minute < 15:
        return True, "pasada diaria de seguridad (06:00 UTC)"

    return False, ""


# ── Bracket (eliminatorias) ───────────────────────────────────────────────────
# stage de football-data → fase interna de la porra
STAGE_MAP: dict[str, str] = {
    "LAST_32":        "ronda_32",
    "LAST_16":        "octavos",
    "QUARTER_FINALS": "cuartos",
    "SEMI_FINALS":    "semifinales",
    "THIRD_PLACE":    "tercer_puesto",
    "FINAL":          "final",
}

# status de football-data → status corto que entiende el dashboard
STATUS_MAP: dict[str, str] = {
    "FINISHED": "FT", "AWARDED": "FT",
    "IN_PLAY": "LIVE", "PAUSED": "LIVE",
    "TIMED": "NS", "SCHEDULED": "NS",
    "POSTPONED": "PST", "SUSPENDED": "SUSP", "CANCELLED": "CANC",
}


def _goles(m: dict) -> tuple[int | None, int | None]:
    ft = (m.get("score") or {}).get("fullTime") or {}
    return ft.get("home"), ft.get("away")


def _build_bracket(all_matches: list[dict]) -> dict[str, list[dict]]:
    """Construye el cuadro eliminatorio a partir de los partidos descargados."""
    bracket: dict[str, list[dict]] = {fase: [] for fase in STAGE_MAP.values()}
    for m in all_matches:
        fase = STAGE_MAP.get(m.get("stage", ""))
        if not fase:
            continue
        gl, gv = _goles(m)
        bracket[fase].append({
            "fixture_id": m.get("id"),
            "date":       m.get("utcDate"),
            "status":     STATUS_MAP.get(m.get("status", ""), m.get("status", "")),
            "team1":      _equipo_es(m.get("homeTeam") or {}),
            "team2":      _equipo_es(m.get("awayTeam") or {}),
            "score1":     gl,
            "score2":     gv,
        })
    for fase, fixes in bracket.items():
        fixes.sort(key=lambda x: x["date"] or "")
        print(f"  {fase}: {len(fixes)} partidos")
    return bracket


def main() -> None:
    if not API_TOKEN:
        print("✘ Variable FOOTBALL_DATA_TOKEN no definida. Saliendo.")
        print("  Regístrate gratis en https://www.football-data.org/client/register")
        sys.exit(1)

    toca, motivo = _toca_llamar()
    if not toca:
        print("▸ Ningún partido ha terminado desde la última actualización — no llamo a la API.")
        return
    print(f"▸ Llamando a la API · motivo: {motivo}")

    print("▸ Descargando partidos del Mundial 2026 (football-data.org, 1 llamada)…")
    all_matches = _get_all_matches()
    print(f"  {len(all_matches)} partidos totales")
    if not all_matches:
        print("✘ La API devolvió 0 partidos. Revisa el token en football-data.org.")
        sys.exit(1)

    terminados = [m for m in all_matches if m.get("status") in TERMINADO]
    print(f"  {len(terminados)} partidos terminados")

    idx = _build_match_index()
    resultados: dict[str, tuple[int, int]] = {}

    for m in terminados:
        home = _equipo_es(m.get("homeTeam") or {})
        away = _equipo_es(m.get("awayTeam") or {})
        gl, gv = _goles(m)
        if gl is None or gv is None:
            continue
        match_id = idx.get((home, away))
        if match_id:
            resultados[match_id] = (int(gl), int(gv))
        else:
            print(f"  ⚠️  Sin mapeo para: {home} vs {away}")

    print(f"  {len(resultados)} partidos mapeados a IDs internos")

    n_grupos = _update_excel(REALES_F, resultados)
    n_elim   = _update_excel(REALES_ELIM, resultados)
    total    = n_grupos + n_elim
    print(f"✓ Excel actualizado — {total} celdas nuevas")

    print("▸ Construyendo cuadro eliminatorio…")
    bracket = _build_bracket(all_matches)
    import datetime as _dt
    BRACKET_F.write_text(
        json.dumps({"updated": _dt.datetime.now(_dt.timezone.utc).isoformat(), "rounds": bracket},
                   ensure_ascii=False, indent=2),
        encoding="utf-8"
    )
    print("✓ bracket.json actualizado")


if __name__ == "__main__":
    main()
