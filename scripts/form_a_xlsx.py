"""
form_a_xlsx.py — Convierte respuestas del Google Form en plantillas xlsx

USO:
    1. Descarga las respuestas: en Google Sheets → Archivo → Descargar → CSV
    2. Ejecuta:
         python scripts/form_a_xlsx.py respuestas.csv

    Genera un archivo pronosticos/plantilla_ALIAS.xlsx por cada participante.

FORMATO DEL FORM:
    Cada partido tiene DOS columnas en el CSV:
      "[G-A1-A2] México 🇲🇽 — Goles"      → goles local
      "[G-A1-A2] 🇿🇦 Sudáfrica — Goles"   → goles visitante
    Los premios tienen columnas: "MVP del Torneo", "Goleador del Torneo", "Portero del Torneo"
"""

import csv
import re
import sys
from pathlib import Path

from openpyxl import load_workbook

# ── Rutas ──────────────────────────────────────────────────────────
RAIZ          = Path(__file__).resolve().parent.parent
DIR_PRONO     = RAIZ / "pronosticos"
PLANTILLA_REF = DIR_PRONO / "plantilla_DaniV.xlsx"


def parsear_columnas_form(headers: list) -> dict:
    """
    Extrae de las cabeceras del CSV los pares (match_id, 'local'/'visit') → nombre_columna.
    Devuelve: { match_id: {'local': col_name, 'visit': col_name} }
    """
    mapa = {}
    for h in headers:
        m = re.match(r'\[([A-Z]-[A-Z0-9]+-[A-Z0-9]+)\].*?—\s*Goles', h)
        if not m:
            continue
        mid = m.group(1)
        if mid not in mapa:
            mapa[mid] = {}
        # Si ya tiene 'local' asignado, esta es 'visit'; si no, es 'local'
        if 'local' not in mapa[mid]:
            mapa[mid]['local'] = h
        else:
            mapa[mid]['visit'] = h
    return mapa


def generar_xlsx(alias: str, apuestas: dict, premios: dict, plantilla_path: Path, destino: Path):
    """
    Copia la plantilla de referencia, escribe goles y premios, guarda en destino.
    apuestas: { match_id: (goles_local, goles_visitante) }
    premios:  { "MVP del Torneo": "...", ... }
    """
    wb = load_workbook(plantilla_path)

    # ── Fase Grupos ────────────────────────────────────────────────
    ws = wb["Fase Grupos"]
    ws["C2"] = alias

    for row in ws.iter_rows(min_row=6, max_row=120):
        mid_cell  = row[2]   # columna C → ID partido
        gl_cell   = row[3]   # columna D → goles local
        gv_cell   = row[4]   # columna E → goles visitante
        if mid_cell.value and str(mid_cell.value).startswith("G-"):
            mid = str(mid_cell.value).strip()
            if mid in apuestas:
                gl_cell.value = apuestas[mid][0]
                gv_cell.value = apuestas[mid][1]

    # ── Premios ────────────────────────────────────────────────────
    ws2 = wb["Premios"]
    FILAS_PREMIOS = {
        "MVP del Torneo":      3,
        "Goleador del Torneo": 4,
        "Portero del Torneo":  5,
    }
    for nombre, fila in FILAS_PREMIOS.items():
        valor = premios.get(nombre, "")
        if valor:
            ws2.cell(row=fila, column=3).value = valor

    wb.save(destino)
    print(f"  ✅  {destino.name}")


def procesar_csv(csv_path: Path):
    with open(csv_path, newline="", encoding="utf-8-sig") as f:
        reader = csv.DictReader(f)
        headers = reader.fieldnames or []
        filas   = list(reader)

    print(f"\n📊 Respuestas: {len(filas)}  |  Columnas: {len(headers)}\n")

    # Mapeo match_id → columnas local/visit
    mapa_partidos = parsear_columnas_form(headers)
    print(f"⚽ Partidos detectados en el CSV: {len(mapa_partidos)}")

    # Columna alias
    col_alias = next(
        (h for h in headers if "alias" in h.lower() or "nombre" in h.lower()),
        headers[1] if len(headers) > 1 else headers[0]
    )
    print(f"🔑 Columna alias: '{col_alias}'\n")

    DIR_PRONO.mkdir(exist_ok=True)

    for fila in filas:
        alias_raw = fila.get(col_alias, "").strip()
        if not alias_raw:
            print("  ⚠️  Fila sin alias, saltando...")
            continue

        alias_safe = re.sub(r'[^\w\-]', '_', alias_raw)

        # Recoger goles por partido
        apuestas = {}
        for mid, cols in mapa_partidos.items():
            col_l = cols.get('local', '')
            col_v = cols.get('visit', '')
            val_l = fila.get(col_l, '').strip()
            val_v = fila.get(col_v, '').strip()
            try:
                apuestas[mid] = (int(val_l), int(val_v))
            except ValueError:
                apuestas[mid] = (None, None)

        # Recoger premios
        premios = {
            k: fila.get(k, '').strip()
            for k in ["MVP del Torneo", "Goleador del Torneo", "Portero del Torneo"]
        }

        rellenos = sum(1 for v in apuestas.values() if v != (None, None))
        print(f"👤 {alias_raw}  →  {rellenos}/{len(mapa_partidos)} partidos rellenos")

        destino = DIR_PRONO / f"plantilla_{alias_safe}.xlsx"
        generar_xlsx(alias_raw, apuestas, premios, PLANTILLA_REF, destino)

    print(f"\n✅ Archivos guardados en: {DIR_PRONO}\n")


if __name__ == "__main__":
    if len(sys.argv) < 2:
        print(__doc__)
        sys.exit(1)
    csv_file = Path(sys.argv[1])
    if not csv_file.exists():
        print(f"❌ No se encuentra: {csv_file}")
        sys.exit(1)
    procesar_csv(csv_file)
