"""
Test E2E del sistema de la porra:
1) Genera 4 Excels de pronóstico ficticios en pronosticos/
2) Genera marcadores reales en data/resultados_reales.xlsx para algunos partidos
3) Ejecuta scorer.py
4) Verifica que los puntos calculados coinciden con la matriz del reglamento:
     - Pleno  (resultado exacto) ............. 3 pts
     - Tendencia (ganador/empate, no exacto) .. 2 pts
     - Fallo ................................. 0 pts

Al final limpia los Excels de prueba para no contaminar la carpeta compartida.
"""

from __future__ import annotations
import json
import shutil
import subprocess
import sys
from pathlib import Path

from openpyxl import load_workbook

RAIZ           = Path(__file__).resolve().parent.parent
DIR_PLANTILLAS = RAIZ / "plantillas"
DIR_PRONO      = RAIZ / "pronosticos"
DIR_DATA       = RAIZ / "data"


def clonar_plantilla(destino: Path):
    shutil.copy(DIR_PLANTILLAS / "plantilla_porra.xlsx", destino)


def rellenar(alias: str, archivo: Path, apuestas_grupos: dict):
    """Rellena alias y apuestas de grupos en la plantilla inicial (sin Eliminatorias)."""
    wb = load_workbook(archivo)
    # El scorer lee el alias sólo de "Fase Grupos" (C2 es la celda top-left de la fusión)
    wb["Fase Grupos"]["C2"] = alias
    ws = wb["Fase Grupos"]
    for fila in ws.iter_rows(min_row=1):
        id_part = fila[2].value
        if isinstance(id_part, str) and id_part in apuestas_grupos:
            gl, gv = apuestas_grupos[id_part]
            fila[3].value = gl
            fila[4].value = gv
    wb.save(archivo)


def rellenar_reales(apuestas: dict):
    archivo = DIR_DATA / "resultados_reales.xlsx"
    wb = load_workbook(archivo)
    ws = wb["Fase Grupos"]
    for fila in ws.iter_rows(min_row=1):
        id_part = fila[2].value
        if isinstance(id_part, str) and id_part in apuestas:
            gl, gv = apuestas[id_part]
            fila[3].value = gl
            fila[4].value = gv
    wb.save(archivo)


def main():
    DIR_PRONO.mkdir(exist_ok=True)
    for f in DIR_PRONO.glob("test_*.xlsx"):
        try:
            f.unlink()
        except (PermissionError, OSError):
            pass  # en algunos sistemas de archivos no se puede borrar; se sobreescribirá

    # ── Resultados reales para 5 partidos del Grupo A ────────────
    REALES = {
        "G-A1-A2": (2, 1),   # México 2-1 Sudáfrica
        "G-A3-A4": (0, 0),   # Corea del Sur 0-0 Chequia
        "G-A1-A3": (3, 0),   # México 3-0 Corea del Sur
        "G-A2-A4": (1, 2),   # Sudáfrica 1-2 Chequia
        "G-A1-A4": (1, 1),   # México 1-1 Chequia
    }
    rellenar_reales(REALES)

    # ── 4 pronósticos con apuestas de grupos ─────────────────────
    #
    # Daniel: 4 plenos → 4×3 = 12 pts
    #   G-A1-A2 (2,1) vs real (2,1) → pleno   3
    #   G-A3-A4 (0,0) vs real (0,0) → pleno   3
    #   G-A1-A3 (3,0) vs real (3,0) → pleno   3
    #   G-A2-A4 (1,2) vs real (1,2) → pleno   3
    #   G-A1-A4 (2,1) vs real (1,1) → fallo   0  (apostó victoria, real empate)
    #
    # Pablo: 4 tendencias → 4×2 = 8 pts
    #   G-A1-A2 (4,2) vs real (2,1) → tend    2
    #   G-A3-A4 (1,1) vs real (0,0) → tend    2
    #   G-A1-A3 (2,0) vs real (3,0) → tend    2
    #   G-A2-A4 (0,0) vs real (1,2) → fallo   0  (apostó empate, real gana visitante)
    #   G-A1-A4 (3,3) vs real (1,1) → tend    2
    #
    # Lola: 1 pleno + 1 tendencia = 3+2 = 5 pts
    #   G-A1-A2 (0,3) vs real (2,1) → fallo   0
    #   G-A3-A4 (2,1) vs real (0,0) → fallo   0
    #   G-A1-A3 (0,1) vs real (3,0) → fallo   0
    #   G-A2-A4 (0,5) vs real (1,2) → tend    2  (gana visitante en ambos)
    #   G-A1-A4 (1,1) vs real (1,1) → pleno   3
    #
    # Marta: no apuesta nada → 0 pts

    PARTICIPANTES = [
        ("Daniel", {
            "G-A1-A2": (2, 1),
            "G-A3-A4": (0, 0),
            "G-A1-A3": (3, 0),
            "G-A2-A4": (1, 2),
            "G-A1-A4": (2, 1),
        }),
        ("Pablo", {
            "G-A1-A2": (4, 2),
            "G-A3-A4": (1, 1),
            "G-A1-A3": (2, 0),
            "G-A2-A4": (0, 0),
            "G-A1-A4": (3, 3),
        }),
        ("Lola", {
            "G-A1-A2": (0, 3),
            "G-A3-A4": (2, 1),
            "G-A1-A3": (0, 1),
            "G-A2-A4": (0, 5),
            "G-A1-A4": (1, 1),
        }),
        ("Marta", {}),
    ]

    archivos_creados = []
    for alias, apuestas in PARTICIPANTES:
        archivo = DIR_PRONO / f"test_{alias.lower()}.xlsx"
        clonar_plantilla(archivo)
        rellenar(alias, archivo, apuestas)
        archivos_creados.append(archivo)

    # ── Ejecutar scorer ──────────────────────────────────────────
    print("▸ Ejecutando scorer.py…")
    out = subprocess.run(
        [sys.executable, str(RAIZ / "scripts" / "scorer.py")],
        capture_output=True, text=True, cwd=RAIZ,
    )
    print(out.stdout)
    if out.returncode != 0:
        print(out.stderr, file=sys.stderr)
        sys.exit(out.returncode)

    # ── Validar resultados ───────────────────────────────────────
    data    = json.loads((RAIZ / "dashboard" / "data.json").read_text(encoding="utf-8"))
    ranking = {r["alias"]: r for r in data["ranking"]}

    # Verificar que el bloque de puntuación en data.json es el correcto
    puntaje = data.get("puntuacion", {})
    assert puntaje.get("pleno")     == 3,  f"PLENO esperado 3, obtenido {puntaje.get('pleno')}"
    assert puntaje.get("tendencia") == 2,  f"TENDENCIA esperada 2, obtenida {puntaje.get('tendencia')}"
    assert puntaje.get("quien_pasa")== 1,  f"QUIEN_PASA esperado 1, obtenido {puntaje.get('quien_pasa')}"
    print("✓ Constantes de puntuación correctas")

    ESPERADO = {
        "Daniel": {"pleno": 4, "tendencia": 0, "puntos_grupos": 12},
        "Pablo":  {"pleno": 0, "tendencia": 4, "puntos_grupos": 8},
        "Lola":   {"pleno": 1, "tendencia": 1, "puntos_grupos": 5},
        "Marta":  {"pleno": 0, "tendencia": 0, "puntos_grupos": 0},
    }

    fallos = 0
    print()
    print(f"{'Alias':<10} {'Pleno':>7} {'Tend.':>7} {'Grupos':>8}  resultado")
    print("─" * 50)
    for alias, esp in ESPERADO.items():
        r = ranking.get(alias)
        if not r:
            print(f"  {alias}: NO ENCONTRADO en ranking"); fallos += 1; continue

        ok = (r["pleno"]         == esp["pleno"]
              and r["tendencia"] == esp["tendencia"]
              and r["puntos_grupos"] == esp["puntos_grupos"])
        marca = "✓" if ok else "✘"
        print(f"{alias:<10} {r['pleno']:>7} {r['tendencia']:>7} {r['puntos_grupos']:>8}  {marca}"
              + ("" if ok else f"  (esperado: pleno={esp['pleno']} tend={esp['tendencia']} grupos={esp['puntos_grupos']})"))
        if not ok:
            fallos += 1

        # Verificar que total = grupos + elim + quien_pasa + especiales (sin campo 'bonus')
        calc = (r.get("puntos_grupos", 0) + r.get("puntos_elim", 0)
                + r.get("puntos_quien_pasa", 0) + r.get("puntos_especiales", 0))
        if calc != r["total"]:
            print(f"  ✘ {alias}: suma de campos ({calc}) != total ({r['total']})")
            fallos += 1

        if "bonus" in r:
            print(f"  ✘ {alias}: campo 'bonus' obsoleto sigue presente en el output")
            fallos += 1

    # Verificar orden del ranking: Daniel > Pablo > Lola > Marta
    print()
    ord_alias = [r["alias"] for r in data["ranking"]]
    print(f"Orden del ranking: {ord_alias}")
    esperado_orden = ["Daniel", "Pablo", "Lola", "Marta"]
    if ord_alias != esperado_orden:
        print(f"✘ Orden esperado: {esperado_orden}")
        fallos += 1
    else:
        print("✓ Orden del ranking correcto")

    # Verificar que no haya bonus de líder de grupos en ningún participante
    if any(r.get("bonus", 0) != 0 for r in data["ranking"]):
        print("✘ Se encontró 'bonus' distinto de 0 (regla eliminada)")
        fallos += 1
    else:
        print("✓ Sin bonus de líder (eliminado correctamente)")

    print()
    if fallos == 0:
        print("══════════════════════════════════════")
        print("  ✔ TODOS LOS TESTS PASAN")
        print("══════════════════════════════════════")
    else:
        print(f"  ✘ {fallos} test(s) fallaron")
        sys.exit(1)

    # ── Limpieza ─────────────────────────────────────────────────
    eliminados = 0
    for f in archivos_creados:
        try:
            f.unlink()
            eliminados += 1
        except (PermissionError, OSError):
            print(f"  (no se pudo eliminar {f.name} — bórralo manualmente)")
    if eliminados:
        print(f"\n({eliminados} Excel(s) de prueba eliminados)")

    # Borrar resultados de prueba de resultados_reales.xlsx
    wb = load_workbook(DIR_DATA / "resultados_reales.xlsx")
    for hoja in wb.sheetnames:
        ws = wb[hoja]
        for fila in ws.iter_rows(min_row=1):
            if isinstance(fila[2].value, str) and (
                fila[2].value.startswith("G-") or fila[2].value.startswith("R")
            ):
                fila[3].value = None
                fila[4].value = None
    wb.save(DIR_DATA / "resultados_reales.xlsx")
    print("(Resultados de prueba borrados de resultados_reales.xlsx)")

    # Regenerar data.json limpio
    subprocess.run([sys.executable, str(RAIZ / "scripts" / "scorer.py")], cwd=RAIZ)


if __name__ == "__main__":
    main()
