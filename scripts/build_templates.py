"""
Genera las plantillas de la porra a partir de data/matches.json.

USO:
    python build_templates.py                   → plantilla FASE DE GRUPOS (por defecto)
    python build_templates.py --fase grupos     → plantilla FASE DE GRUPOS
    python build_templates.py --fase elim       → plantilla ELIMINATORIAS

PLANTILLA DE GRUPOS  (plantillas/plantilla_porra.xlsx)
    Hojas: README · Fase Grupos · Premios
    Se distribuye ANTES del torneo.
    Nombre de archivo del participante: porra_TUNOMBRE.xlsx

PLANTILLA DE ELIMINATORIAS  (plantillas/plantilla_eliminatorias.xlsx)
    Hojas: README · Eliminatorias
    Se distribuye CUANDO SE CONOCEN los emparejamientos (tras la fase de grupos).
    Nombre de archivo del participante: porra_TUNOMBRE_elim.xlsx
    IMPORTANTE: usar el MISMO ALIAS que en la plantilla de grupos.

Estructura del Excel:
    Celda C2  → Alias
    Columna C → ID de Partido (inmutable)
    Columna D → Goles equipo local
    Columna E → Goles equipo visitante
    Columna G → Quien Pasa (solo hoja Eliminatorias)
"""

import json
import sys
from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.worksheet.datavalidation import DataValidation


# ───────────────────────── estilos ──────────────────────────
FUENTE_BASE    = "Calibri"
COLOR_AZUL     = "1F4E78"
COLOR_AZUL_BG  = "D9E1F2"
COLOR_GRIS_BG  = "F2F2F2"
COLOR_AMBAR    = "FFE699"
COLOR_ROJO_BG  = "FCE4D6"
COLOR_VERDE_BG = "E2EFDA"

borde_fino = Border(
    left=Side(style="thin", color="BFBFBF"),
    right=Side(style="thin", color="BFBFBF"),
    top=Side(style="thin", color="BFBFBF"),
    bottom=Side(style="thin", color="BFBFBF"),
)

_BANDERAS = {}
try:
    _ruta_flags = Path(__file__).resolve().parent.parent / "data" / "flags.json"
    if _ruta_flags.exists():
        _BANDERAS = json.loads(_ruta_flags.read_text(encoding="utf-8"))
except Exception:
    _BANDERAS = {}


def con_bandera(nombre, lado="local"):
    flag = _BANDERAS.get(nombre, "")
    if not flag:
        return nombre
    return f"{flag} {nombre}" if lado == "visitante" else f"{nombre} {flag}"


# ───────────────────────── helpers ──────────────────────────
def estilo_titulo(celda):
    celda.font = Font(name=FUENTE_BASE, size=14, bold=True, color="FFFFFF")
    celda.fill = PatternFill("solid", start_color=COLOR_AZUL)
    celda.alignment = Alignment(horizontal="center", vertical="center")


def estilo_cabecera(celda):
    celda.font = Font(name=FUENTE_BASE, size=11, bold=True, color=COLOR_AZUL)
    celda.fill = PatternFill("solid", start_color=COLOR_AZUL_BG)
    celda.alignment = Alignment(horizontal="center", vertical="center")
    celda.border = borde_fino


def celda_alias(ws, n_cols, alias_default="", resultados=False):
    """Fila 2: etiqueta + celda de alias."""
    ws["B2"] = "Alias / Nombre de Apuesta:"
    ws["B2"].font = Font(name=FUENTE_BASE, size=11, bold=True, color=COLOR_AZUL)
    ws["B2"].alignment = Alignment(horizontal="right", vertical="center")
    ultima = chr(ord("B") + n_cols)
    ws.merge_cells(f"C2:{ultima}2")
    ca = ws["C2"]
    ca.value = alias_default if alias_default else "AdminOficina" if resultados else ""
    ca.font  = Font(name=FUENTE_BASE, size=12, bold=True)
    ca.fill  = PatternFill("solid", start_color=COLOR_AMBAR)
    ca.alignment = Alignment(horizontal="left", vertical="center", indent=1)
    ca.border = borde_fino
    ws.row_dimensions[2].height = 22


# ──────────────────────── hoja grupos ───────────────────────
def construir_hoja_grupos(ws, partidos, *, alias_default="", resultados=False):
    ws.sheet_view.showGridLines = False
    for c, w in {"A":3,"B":22,"C":28,"D":9,"E":9,"F":28,"G":18}.items():
        ws.column_dimensions[c].width = w

    ws.merge_cells("B1:G1")
    ws["B1"] = "FASE DE GRUPOS  ·  72 partidos" + ("   ·   RESULTADOS REALES" if resultados else "")
    estilo_titulo(ws["B1"])
    ws.row_dimensions[1].height = 26

    celda_alias(ws, 5, alias_default, resultados)

    ws.merge_cells("B3:G3")
    av = ws["B3"]
    av.value = ("Rellena los goles previstos en cada partido. "
                "Fecha limite: 11 jun 2026 a las 17:00 h.")
    av.font  = Font(name=FUENTE_BASE, size=10, italic=True, color="3F3F3F")
    av.fill  = PatternFill("solid", start_color=COLOR_GRIS_BG)
    av.alignment = Alignment(horizontal="center", vertical="center")
    av.border = borde_fino
    ws.row_dimensions[3].height = 20

    CABECERAS = ["ID Partido","Equipo Local","Goles L","Goles V","Equipo Visitante","Fase / Grupo"]
    for idx, txt in enumerate(CABECERAS, start=2):
        estilo_cabecera(ws.cell(row=5, column=idx, value=txt))
    ws.row_dimensions[5].height = 22

    dv = DataValidation(type="whole", operator="between", formula1=0, formula2=20,
                        showErrorMessage=True, errorTitle="Valor no valido",
                        error="Introduce un numero entero entre 0 y 20.")
    ws.add_data_validation(dv)

    fila = 6
    grupo_actual = None
    for p in partidos:
        if p.get("grupo") != grupo_actual:
            grupo_actual = p["grupo"]
            ws.merge_cells(start_row=fila, start_column=2, end_row=fila, end_column=7)
            sep = ws.cell(row=fila, column=2, value=f"  Grupo {grupo_actual}")
            sep.font = Font(name=FUENTE_BASE, size=11, bold=True, color=COLOR_AZUL)
            sep.fill = PatternFill("solid", start_color=COLOR_GRIS_BG)
            sep.alignment = Alignment(horizontal="left", vertical="center")
            ws.row_dimensions[fila].height = 18
            fila += 1

        ws.cell(row=fila, column=3, value=p["id"]).font = Font(
            name="Consolas", size=10, bold=True, color="595959")
        ws.cell(row=fila, column=3).alignment = Alignment(horizontal="center", vertical="center")

        c_loc = ws.cell(row=fila, column=2, value=con_bandera(p["local"], "local"))
        c_loc.font = Font(name=FUENTE_BASE, size=11, bold=True)
        c_loc.alignment = Alignment(horizontal="right", vertical="center", indent=1)

        for col in (4, 5):
            c = ws.cell(row=fila, column=col)
            c.font = Font(name=FUENTE_BASE, size=12, bold=True, color="1F4E78")
            c.alignment = Alignment(horizontal="center", vertical="center")
            c.fill = PatternFill("solid", start_color=COLOR_AMBAR)
            dv.add(c)

        c_vis = ws.cell(row=fila, column=6, value=con_bandera(p["visitante"], "visitante"))
        c_vis.font = Font(name=FUENTE_BASE, size=11, bold=True)
        c_vis.alignment = Alignment(horizontal="left", vertical="center", indent=1)

        c_fase = ws.cell(row=fila, column=7,
                         value=f"Grupo {p['grupo']} · J{p['jornada']}")
        c_fase.font = Font(name=FUENTE_BASE, size=9, italic=True, color="7F7F7F")
        c_fase.alignment = Alignment(horizontal="center", vertical="center")

        for col in range(2, 8):
            ws.cell(row=fila, column=col).border = borde_fino
        fila += 1

    ws.freeze_panes = "A6"


# ──────────────────────── hoja eliminatorias ────────────────
def construir_hoja_elim(ws, partidos, *, alias_default="", resultados=False):
    ws.sheet_view.showGridLines = False
    for c, w in {"A":3,"B":22,"C":28,"D":9,"E":9,"F":22,"G":22,"H":18}.items():
        ws.column_dimensions[c].width = w

    ws.merge_cells("B1:H1")
    ws["B1"] = ("FASE ELIMINATORIA  ·  32 partidos  ·  resultado al min. 90 + descuento"
                + ("   ·   RESULTADOS REALES" if resultados else ""))
    estilo_titulo(ws["B1"])
    ws.row_dimensions[1].height = 26

    celda_alias(ws, 6, alias_default, resultados)

    ws.merge_cells("B3:H3")
    av = ws["B3"]
    av.value = ("REGLA DE ORO: anota el marcador al termino de los 90' + descuento (sin prorroga ni penaltis). "
                "Columna G (verde): equipo que crees que pasa de ronda (+1 pt). "
                "USA EL MISMO ALIAS que en tu archivo de grupos.")
    av.font  = Font(name=FUENTE_BASE, size=10, italic=True, color="9C0006")
    av.fill  = PatternFill("solid", start_color=COLOR_ROJO_BG)
    av.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    av.border = borde_fino
    ws.row_dimensions[3].height = 34

    CABECERAS = ["ID Partido","Equipo Local","Goles L","Goles V","Equipo Visitante","Quien Pasa","Fase / Ronda"]
    for idx, txt in enumerate(CABECERAS, start=2):
        estilo_cabecera(ws.cell(row=5, column=idx, value=txt))
    # Cabecera "Quien Pasa" en verde oscuro
    ws.cell(row=5, column=7).fill = PatternFill("solid", start_color="375623")
    ws.cell(row=5, column=7).font = Font(name=FUENTE_BASE, size=11, bold=True, color="FFFFFF")
    ws.row_dimensions[5].height = 22

    dv = DataValidation(type="whole", operator="between", formula1=0, formula2=20,
                        showErrorMessage=True, errorTitle="Valor no valido",
                        error="Introduce un numero entero entre 0 y 20.")
    ws.add_data_validation(dv)

    NOMBRE_FASE = {
        "ronda_32":"Ronda de 32","octavos":"Octavos","cuartos":"Cuartos",
        "semifinales":"Semifinales","tercer_puesto":"Tercer Puesto","final":"Final",
    }

    fila = 6
    fase_actual = None
    for p in partidos:
        if p.get("fase") != fase_actual:
            fase_actual = p["fase"]
            ws.merge_cells(start_row=fila, start_column=2, end_row=fila, end_column=8)
            sep = ws.cell(row=fila, column=2,
                          value=f"  {NOMBRE_FASE.get(fase_actual, fase_actual.upper())}")
            sep.font = Font(name=FUENTE_BASE, size=11, bold=True, color=COLOR_AZUL)
            sep.fill = PatternFill("solid", start_color=COLOR_GRIS_BG)
            sep.alignment = Alignment(horizontal="left", vertical="center")
            ws.row_dimensions[fila].height = 18
            fila += 1

        ws.cell(row=fila, column=3, value=p["id"]).font = Font(
            name="Consolas", size=10, bold=True, color="595959")
        ws.cell(row=fila, column=3).alignment = Alignment(horizontal="center", vertical="center")

        c_loc = ws.cell(row=fila, column=2, value=con_bandera(p["local"], "local"))
        c_loc.font = Font(name=FUENTE_BASE, size=11, bold=True)
        c_loc.alignment = Alignment(horizontal="right", vertical="center", indent=1)

        for col in (4, 5):
            c = ws.cell(row=fila, column=col)
            c.font = Font(name=FUENTE_BASE, size=12, bold=True, color="1F4E78")
            c.alignment = Alignment(horizontal="center", vertical="center")
            c.fill = PatternFill("solid", start_color=COLOR_AMBAR)
            dv.add(c)

        c_vis = ws.cell(row=fila, column=6, value=con_bandera(p["visitante"], "visitante"))
        c_vis.font = Font(name=FUENTE_BASE, size=11, bold=True)
        c_vis.alignment = Alignment(horizontal="left", vertical="center", indent=1)

        c_qp = ws.cell(row=fila, column=7)
        c_qp.font = Font(name=FUENTE_BASE, size=11, bold=True, color="375623")
        c_qp.fill = PatternFill("solid", start_color=COLOR_VERDE_BG)
        c_qp.alignment = Alignment(horizontal="center", vertical="center")

        c_fase = ws.cell(row=fila, column=8,
                         value=NOMBRE_FASE.get(p.get("fase",""), p.get("fase","")))
        c_fase.font = Font(name=FUENTE_BASE, size=9, italic=True, color="7F7F7F")
        c_fase.alignment = Alignment(horizontal="center", vertical="center")

        for col in range(2, 9):
            ws.cell(row=fila, column=col).border = borde_fino
        fila += 1

    ws.freeze_panes = "A6"


# ──────────────────────── hoja premios ──────────────────────
def construir_hoja_premios(ws):
    ws.sheet_view.showGridLines = False
    for c, w in {"A":3,"B":28,"C":35,"D":20}.items():
        ws.column_dimensions[c].width = w

    ws.merge_cells("B1:D1")
    ws["B1"] = "PREMIOS ESPECIALES  ·  +15 pts cada uno"
    estilo_titulo(ws["B1"])
    ws.row_dimensions[1].height = 26

    ws.merge_cells("B2:D2")
    av = ws["B2"]
    av.value = ("Escribe el nombre del jugador que crees ganara cada premio. "
                "Fecha limite: 11 jun 2026 a las 17:00 h.")
    av.font  = Font(name=FUENTE_BASE, size=10, italic=True, color="3F3F3F")
    av.fill  = PatternFill("solid", start_color=COLOR_GRIS_BG)
    av.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    av.border = borde_fino
    ws.row_dimensions[2].height = 28

    for i, (premio, desc) in enumerate([
        ("MVP del Torneo",      "Mejor jugador del torneo (FIFA)"),
        ("Goleador del Torneo", "Maximo anotador · Bota de Oro FIFA"),
        ("Portero del Torneo",  "Mejor portero · Guante de Oro FIFA"),
    ], start=3):
        c_p = ws.cell(row=i, column=2, value=premio)
        c_p.font = Font(name=FUENTE_BASE, size=12, bold=True, color=COLOR_AZUL)
        c_p.alignment = Alignment(horizontal="right", vertical="center", indent=1)
        c_p.border = borde_fino

        c_i = ws.cell(row=i, column=3)
        c_i.font = Font(name=FUENTE_BASE, size=12, bold=True)
        c_i.fill = PatternFill("solid", start_color=COLOR_AMBAR)
        c_i.alignment = Alignment(horizontal="left", vertical="center", indent=1)
        c_i.border = borde_fino

        c_d = ws.cell(row=i, column=4, value=desc)
        c_d.font = Font(name=FUENTE_BASE, size=9, italic=True, color="7F7F7F")
        c_d.alignment = Alignment(horizontal="left", vertical="center", indent=1)
        c_d.border = borde_fino
        ws.row_dimensions[i].height = 26


# ──────────────────────── README ────────────────────────────
def construir_readme_grupos(ws):
    ws.sheet_view.showGridLines = False
    ws.column_dimensions["A"].width = 2
    ws.column_dimensions["B"].width = 110
    ws["B2"] = "PORRA DEL MUNDIAL 2026  —  PLANTILLA FASE DE GRUPOS"
    estilo_titulo(ws["B2"])
    ws.row_dimensions[2].height = 28
    lines = [
        "1. Escribe tu ALIAS en la celda C2 de las hojas 'Fase Grupos' y 'Premios'. "
            "Recuerda este alias: lo necesitaras en la plantilla de eliminatorias.",
        "2. Rellena los goles previstos en las columnas amarillas (D = local, E = visitante).",
        "3. En la hoja 'Premios', escribe el nombre del jugador que crees ganara "
            "el MVP, el Goleador y el Portero del torneo.",
        "4. Guarda el archivo como  porra_TUNOMBRE.xlsx  y dejalo en la carpeta 'pronosticos/'.",
        "5. Fecha limite: 11 junio 2026 a las 17:00 h (hora Espana).",
        "6. Puntuacion: Pleno (goles exactos) = 3 pts | Tendencia (ganador/empate) = 2 pts | Fallo = 0 pts.",
        "7. Tras la fase de grupos recibiras una segunda plantilla para las eliminatorias "
            "(porra_TUNOMBRE_elim.xlsx). Usa el MISMO ALIAS.",
    ]
    for i, txt in enumerate(lines, start=4):
        c = ws.cell(row=i, column=2, value=txt)
        c.font = Font(name=FUENTE_BASE, size=11)
        c.alignment = Alignment(wrap_text=True, vertical="center")
        ws.row_dimensions[i].height = 28


def construir_readme_elim(ws):
    ws.sheet_view.showGridLines = False
    ws.column_dimensions["A"].width = 2
    ws.column_dimensions["B"].width = 110
    ws["B2"] = "PORRA DEL MUNDIAL 2026  —  PLANTILLA ELIMINATORIAS"
    estilo_titulo(ws["B2"])
    ws.row_dimensions[2].height = 28
    lines = [
        "IMPORTANTE: usa el MISMO ALIAS que pusiste en tu plantilla de grupos (porra_TUNOMBRE.xlsx).",
        "1. Escribe tu alias en la celda C2 de la hoja 'Eliminatorias'.",
        "2. Rellena los goles al termino de los 90' + descuento (columnas D y E, amarillo). "
            "SIN prorroga ni penaltis.",
        "3. En la columna G (verde 'Quien Pasa') escribe el equipo que crees que avanza de ronda (+1 pt). "
            "Relevante si predices un empate a 90'.",
        "4. Guarda el archivo como  porra_TUNOMBRE_elim.xlsx  (con el sufijo _elim) "
            "en la carpeta 'pronosticos/'.",
        "5. Fechas limite:  R32: 28 jun 17:00h | Octavos: 4 jul 17:00h | Cuartos: 9 jul 17:00h | "
            "Semis: 14 jul 17:00h | Final: 17 jul 17:00h.",
    ]
    for i, txt in enumerate(lines, start=4):
        c = ws.cell(row=i, column=2, value=txt)
        c.font = Font(name=FUENTE_BASE, size=11)
        c.alignment = Alignment(wrap_text=True, vertical="center")
        ws.row_dimensions[i].height = 28


# ──────────────────────── entrypoints ───────────────────────
def generar_plantilla_grupos(aqui, partidos):
    grupos = [p for p in partidos if p["fase"] == "grupos"]
    assert len(grupos) == 72

    wb = Workbook()
    ws_readme = wb.active; ws_readme.title = "README"
    construir_readme_grupos(ws_readme)
    ws_g = wb.create_sheet("Fase Grupos")
    construir_hoja_grupos(ws_g, grupos)
    ws_p = wb.create_sheet("Premios")
    construir_hoja_premios(ws_p)

    salida = aqui / "plantillas" / "plantilla_porra.xlsx"
    salida.parent.mkdir(parents=True, exist_ok=True)
    wb.save(salida)
    print(f"OK · plantilla grupos → {salida}")

    # Resultados reales (admin) — solo fase de grupos
    wb_r = Workbook()
    wr_g = wb_r.active; wr_g.title = "Fase Grupos"
    construir_hoja_grupos(wr_g, grupos, alias_default="AdminOficina", resultados=True)
    wb_r.save(aqui / "data" / "resultados_reales.xlsx")
    print(f"OK · resultados reales grupos → {aqui / 'data' / 'resultados_reales.xlsx'}")


def generar_plantilla_elim(aqui, partidos):
    elim = [p for p in partidos if p["fase"] != "grupos"]
    assert len(elim) == 32

    wb = Workbook()
    ws_readme = wb.active; ws_readme.title = "README"
    construir_readme_elim(ws_readme)
    ws_e = wb.create_sheet("Eliminatorias")
    construir_hoja_elim(ws_e, elim)

    salida = aqui / "plantillas" / "plantilla_eliminatorias.xlsx"
    salida.parent.mkdir(parents=True, exist_ok=True)
    wb.save(salida)
    print(f"OK · plantilla eliminatorias → {salida}")

    # Resultados reales eliminatorias (admin)
    wb_r = Workbook()
    wr_e = wb_r.active; wr_e.title = "Eliminatorias"
    construir_hoja_elim(wr_e, elim, alias_default="AdminOficina", resultados=True)
    dest = aqui / "data" / "resultados_reales_elim.xlsx"
    wb_r.save(dest)
    print(f"OK · resultados reales elim → {dest}")

    # JSON de apoyo (solo si no existen)
    for nombre, contenido in [
        ("premios_reales.json",
         {"_comentario": "Rellena al finalizar el torneo.",
          "mvp": "", "goleador": "", "portero": ""}),
        ("quien_paso.json",
         {"_comentario": "Solo para partidos que acabaron en empate a 90'. "
                         "Ej: {\"R32-01\": \"Espana\"}"}),
    ]:
        p = aqui / "data" / nombre
        if not p.exists():
            p.write_text(json.dumps(contenido, ensure_ascii=False, indent=2), encoding="utf-8")
            print(f"OK · {nombre} → {p}")


def main():
    fase = "grupos"
    if "--fase" in sys.argv:
        idx = sys.argv.index("--fase")
        if idx + 1 < len(sys.argv):
            fase = sys.argv[idx + 1].lower()
    # shorthand: python build_templates.py elim
    for arg in sys.argv[1:]:
        if arg in ("elim", "eliminatorias"):
            fase = "elim"

    aqui = Path(__file__).resolve().parent.parent
    datos = json.loads((aqui / "data" / "matches.json").read_text(encoding="utf-8"))
    partidos = datos["partidos"]

    if fase in ("grupos", "grupos"):
        generar_plantilla_grupos(aqui, partidos)
    elif fase in ("elim", "eliminatorias"):
        generar_plantilla_elim(aqui, partidos)
    else:
        print(f"Fase no reconocida: '{fase}'. Usa 'grupos' o 'elim'.")
        sys.exit(1)


if __name__ == "__main__":
    main()
