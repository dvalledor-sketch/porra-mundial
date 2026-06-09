"""
supabase_pull.py — Descarga los pronósticos enviados desde el formulario web
(Supabase) y los convierte en archivos pronosticos/porra_ALIAS.xlsx que el
scorer ya sabe leer. El scorer NO se modifica.

Respeta la regla de la hora límite: el mtime de cada archivo generado se fija
al updated_at del participante en la base de datos (la hora oficial de registro).

USO (admin / CI):
    export SUPABASE_URL="https://iwicyguwsobdbsusxydx.supabase.co"
    export SUPABASE_SERVICE_KEY="<service_role key>"   # NO el publishable
    python scripts/supabase_pull.py

El service_role key se guarda como secret (GitHub Actions) o variable de entorno
local. Nunca se publica en el sitio web.
"""

from __future__ import annotations

import json
import os
import re
import sys
import urllib.request
import urllib.error
from datetime import datetime, timezone
from pathlib import Path

from openpyxl import load_workbook

# ── Rutas ────────────────────────────────────────────────────────
RAIZ        = Path(__file__).resolve().parent.parent
DIR_PRONO   = RAIZ / "pronosticos"
PLANTILLA   = RAIZ / "pronosticos" / "plantillas" / "plantilla_porra.xlsx"

FILAS_PREMIOS = {"MVP del Torneo": 3, "Goleador del Torneo": 4, "Portero del Torneo": 5}


# ── Generador de xlsx (compatible con scorer.py) ─────────────────
def generar_xlsx(alias: str, apuestas: dict, premios: dict, destino: Path):
    """apuestas: {match_id: (gl, gv)}  ·  premios: {'mvp':..,'goleador':..,'portero':..}"""
    wb = load_workbook(PLANTILLA)

    ws = wb["Fase Grupos"]
    ws["C2"] = alias
    for row in ws.iter_rows(min_row=6, max_row=120):
        mid_cell, gl_cell, gv_cell = row[2], row[3], row[4]   # C, D, E
        if mid_cell.value and str(mid_cell.value).startswith("G-"):
            mid = str(mid_cell.value).strip()
            if mid in apuestas:
                gl_cell.value, gv_cell.value = apuestas[mid]

    ws2 = wb["Premios"]
    mapa = {
        "MVP del Torneo":      premios.get("mvp"),
        "Goleador del Torneo": premios.get("goleador"),
        "Portero del Torneo":  premios.get("portero"),
    }
    for nombre, fila in FILAS_PREMIOS.items():
        if mapa.get(nombre):
            ws2.cell(row=fila, column=3).value = mapa[nombre]

    wb.save(destino)


def escribir_todos(participantes, grupos, premios, dir_out: Path = DIR_PRONO) -> int:
    """
    participantes: [{alias_norm, alias, updated_at}]
    grupos:        [{alias_norm, match_id, goles_local, goles_visitante}]
    premios:       [{alias_norm, mvp, goleador, portero}]
    Devuelve nº de archivos generados.
    """
    dir_out.mkdir(exist_ok=True)
    g_por_alias, p_por_alias = {}, {}
    for g in grupos:
        g_por_alias.setdefault(g["alias_norm"], {})[g["match_id"]] = (
            g["goles_local"], g["goles_visitante"])
    for p in premios:
        p_por_alias[p["alias_norm"]] = p

    n = 0
    for part in participantes:
        an = part["alias_norm"]
        apuestas = g_por_alias.get(an, {})
        prem = p_por_alias.get(an, {})
        if not apuestas and not any(prem.get(k) for k in ("mvp", "goleador", "portero")):
            continue  # participante sin nada que puntuar todavía

        alias_safe = re.sub(r"[^\w\-]", "_", part["alias"])
        destino = dir_out / f"porra_{alias_safe}.xlsx"
        generar_xlsx(part["alias"], apuestas, prem, destino)

        # Fijar mtime = updated_at (hora oficial de registro)
        try:
            ts = datetime.fromisoformat(part["updated_at"].replace("Z", "+00:00"))
            epoch = ts.timestamp()
            os.utime(destino, (epoch, epoch))
        except Exception as e:
            print(f"  ⚠ no se pudo fijar mtime de {destino.name}: {e}")

        n += 1
        print(f"  ✅ {destino.name}  ({len(apuestas)}/72 partidos)")
    return n


# ── Acceso REST a Supabase (service_role) ────────────────────────
def _get(url: str, key: str):
    req = urllib.request.Request(
        url,
        headers={"apikey": key, "Authorization": f"Bearer {key}",
                 "Accept": "application/json"},
    )
    with urllib.request.urlopen(req, timeout=30) as r:
        return json.loads(r.read().decode("utf-8"))


def fetch_supabase(base_url: str, key: str):
    rest = base_url.rstrip("/") + "/rest/v1"
    participantes = _get(f"{rest}/porra_participantes?select=alias_norm,alias,updated_at", key)
    grupos        = _get(f"{rest}/porra_pronosticos_grupos?select=alias_norm,match_id,goles_local,goles_visitante", key)
    premios       = _get(f"{rest}/porra_premios?select=alias_norm,mvp,goleador,portero", key)
    return participantes, grupos, premios


def main():
    base = os.environ.get("SUPABASE_URL")
    key  = os.environ.get("SUPABASE_SERVICE_KEY")
    if not base or not key:
        print("❌ Falta SUPABASE_URL o SUPABASE_SERVICE_KEY en el entorno.")
        sys.exit(1)
    print("⬇  Descargando pronósticos de Supabase…")
    try:
        parts, grupos, premios = fetch_supabase(base, key)
    except urllib.error.HTTPError as e:
        print(f"❌ HTTP {e.code}: {e.read().decode('utf-8', 'ignore')}")
        sys.exit(1)
    print(f"   {len(parts)} participantes · {len(grupos)} filas de grupos")
    n = escribir_todos(parts, grupos, premios)
    print(f"✅ {n} archivos generados en {DIR_PRONO}")


if __name__ == "__main__":
    main()
