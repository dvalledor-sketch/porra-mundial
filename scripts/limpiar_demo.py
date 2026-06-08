"""
Borra los datos de DEMO y deja el sistema en blanco para producción:
- elimina pronosticos/demo_*.xlsx
- limpia las celdas D y E de data/resultados_reales.xlsx
- regenera dashboard/data.json vacío
"""
import subprocess
import sys
from pathlib import Path
from openpyxl import load_workbook

RAIZ = Path(__file__).resolve().parent.parent
DIR_PRONO = RAIZ / "pronosticos"
DIR_DATA  = RAIZ / "data"


def main():
    # 1) Borrar pronósticos de demo
    borrados = 0
    for f in DIR_PRONO.glob("demo_*.xlsx"):
        try:
            f.unlink()
            borrados += 1
        except PermissionError:
            print(f"⚠ No se pudo borrar {f.name} — bórralo manualmente.", file=sys.stderr)
    print(f"  · {borrados} pronósticos demo eliminados")

    # 2) Limpiar resultados reales
    archivo = DIR_DATA / "resultados_reales.xlsx"
    wb = load_workbook(archivo)
    for hoja in wb.sheetnames:
        if hoja not in ("Fase Grupos", "Eliminatorias"):
            continue
        ws = wb[hoja]
        for fila in ws.iter_rows(min_row=1):
            id_part = fila[2].value
            if isinstance(id_part, str) and (id_part.startswith("G-")
                                             or id_part.startswith("R")
                                             or id_part in ("SF-01","SF-02","TER","FIN","QF-01","QF-02","QF-03","QF-04")):
                fila[3].value = None
                fila[4].value = None
    wb.save(archivo)
    print("  · resultados_reales.xlsx limpiado")

    # 3) Regenerar data.json vacío
    subprocess.run([sys.executable, str(RAIZ / "scripts" / "scorer.py")], cwd=RAIZ, check=True)
    print()
    print("✔ Sistema en blanco. Listo para producción.")


if __name__ == "__main__":
    main()
