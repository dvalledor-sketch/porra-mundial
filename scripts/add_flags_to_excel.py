"""
Añade la bandera (emoji) al lado del nombre de cada selección en los Excel ya
existentes, SIN tocar los goles ni el resto del contenido.

- Columna B (Equipo Local)      → "Nombre 🏳"
- Columna F (Equipo Visitante)  → "🏳 Nombre"

Es idempotente: si la celda ya tiene bandera, no la duplica.
Recorre plantillas/plantilla_porra.xlsx y data/resultados_reales.xlsx.
"""

import json
from pathlib import Path

from openpyxl import load_workbook

RAIZ = Path(__file__).resolve().parent.parent
BANDERAS = json.loads((RAIZ / "data" / "flags.json").read_text(encoding="utf-8"))
TODAS_LAS_FLAGS = set(BANDERAS.values())

ARCHIVOS = [
    RAIZ / "plantillas" / "plantilla_porra.xlsx",
    RAIZ / "data" / "resultados_reales.xlsx",
]

HOJAS = {"Fase Grupos", "Eliminatorias"}


def ya_tiene_bandera(texto: str) -> bool:
    return any(f in texto for f in TODAS_LAS_FLAGS)


def patch(path: Path) -> int:
    if not path.exists():
        print(f"⚠ No existe {path.name}, se omite")
        return 0
    wb = load_workbook(path)
    cambios = 0
    for hoja in wb.sheetnames:
        if hoja not in HOJAS:
            continue
        ws = wb[hoja]
        for fila in ws.iter_rows():
            for celda in fila:
                v = celda.value
                if not isinstance(v, str):
                    continue
                nombre = v.strip()
                if nombre not in BANDERAS or ya_tiene_bandera(v):
                    continue
                flag = BANDERAS[nombre]
                # Columna B (2) = local → bandera a la derecha; F (6) = visitante → izquierda
                if celda.column == 6:
                    celda.value = f"{flag} {nombre}"
                else:
                    celda.value = f"{nombre} {flag}"
                cambios += 1
    if cambios:
        wb.save(path)
    print(f"OK · {path.name}: {cambios} equipos con bandera")
    return cambios


def main():
    total = 0
    for a in ARCHIVOS:
        total += patch(a)
    print(f"Total: {total} celdas actualizadas")


if __name__ == "__main__":
    main()
