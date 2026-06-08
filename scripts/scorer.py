"""
Motor de cálculo de la Porra del Mundial 2026.
─────────────────────────────────────────────
Lee:
    pronosticos/*.xlsx          ← plantillas rellenadas por cada participante
    data/resultados_reales.xlsx ← marcadores oficiales que rellena el admin
    data/matches.json           ← partidos del torneo
    data/premios_reales.json    ← MVP, Goleador y Portero reales (lo rellena el admin al final)
    data/quien_paso.json        ← equipo que pasó en partidos que acabaron en empate a 90'

Aplica el reglamento:
    Pleno absoluto (resultado exacto) ............... 3 pts
    Acierto de tendencia (ganador/empate a 90')...... 2 pts
    Fallo ........................................... 0 pts
    Quién pasa (solo eliminatorias) ................. +1 pt
    MVP del Torneo .................................. +15 pts
    Goleador del Torneo ............................. +15 pts
    Portero del Torneo .............................. +15 pts

Reglas especiales:
    • Una sola entrega por alias (vale la más reciente por modifiedTime).
    • Si el archivo se modificó DESPUÉS de la hora límite de su fase → se invalida
      esa fase en concreto. Las horas límite se definen en data/limites.json.
    • Eliminatorias: marcador al min 90 + descuento (no prórroga, no penaltis).
    • Premios especiales: se pronostican antes de la fecha límite de grupos.

Escribe:
    dashboard/data.json   ← consumido por index.html
"""

from __future__ import annotations

import json
import os
import sys
from dataclasses import dataclass, field
from datetime import datetime, timezone
from pathlib import Path
from typing import Dict, List, Optional

from openpyxl import load_workbook


# ────────────────────────────── rutas ──────────────────────────────
RAIZ              = Path(__file__).resolve().parent.parent
DIR_PRONO         = RAIZ / "pronosticos"
DIR_DATA          = RAIZ / "data"
DIR_DASHBOARD     = RAIZ / "dashboard"
ARCHIVO_PARTIDOS  = DIR_DATA / "matches.json"
ARCHIVO_REALES    = DIR_DATA / "resultados_reales.xlsx"
ARCHIVO_LIMITES   = DIR_DATA / "limites.json"
ARCHIVO_BANDERAS  = DIR_DATA / "flags.json"
ARCHIVO_REALES_ELIM = DIR_DATA / "resultados_reales_elim.xlsx"
ARCHIVO_PREMIOS_R   = DIR_DATA / "premios_reales.json"
ARCHIVO_QUIEN_P     = DIR_DATA / "quien_paso.json"
ARCHIVO_SALIDA      = DIR_DASHBOARD / "data.json"
ARCHIVO_AVATARES    = DIR_DATA / "avatares.json"

PUNTOS_PLENO          = 3
PUNTOS_TENDENCIA      = 2
PUNTOS_QUIEN_PASA     = 1
PUNTOS_BONUS_ESPECIAL = 15

FASES_ELIMINATORIAS = {"ronda_32", "octavos", "cuartos", "semifinales", "tercer_puesto", "final"}
ENCABEZADOS_TABLA   = ["ID Partido", "Equipo Local", "Goles L", "Goles V", "Equipo Visitante"]


# ────────────────────────── modelos ────────────────────────────────
@dataclass
class Pronostico:
    alias: str
    archivo: str
    modified_time: datetime
    apuestas: Dict[str, tuple]          # id_partido -> (gl, gv)
    quien_pasa: Dict[str, str]          # id_partido -> nombre_equipo (solo eliminatorias)
    premios_especiales: Dict[str, str]  # "mvp"/"goleador"/"portero" -> nombre_jugador
    valido_grupos: bool = True
    valido_elim:   bool = True


@dataclass
class Resultado:
    real: Optional[tuple] = None  # (gl, gv) o None si aún no se ha jugado


@dataclass
class DetalleAcierto:
    id_partido: str
    apostado: Optional[tuple]
    real:     Optional[tuple]
    puntos:   int
    tipo:     str  # "pleno" | "tendencia" | "fallo" | "pendiente" | "no_apostado"
    quien_pasa_apostado: Optional[str] = None
    quien_pasa_real:     Optional[str] = None
    puntos_quien_pasa:   int = 0


@dataclass
class Participante:
    alias: str
    archivo: str
    modified_time: datetime
    puntos_grupos:     int = 0
    puntos_elim:       int = 0
    puntos_quien_pasa: int = 0
    puntos_especiales: int = 0
    detalle: List[DetalleAcierto] = field(default_factory=list)
    aciertos_pleno:      int = 0
    aciertos_tendencia:  int = 0
    aciertos_quien_pasa: int = 0
    valido_grupos: bool = True
    valido_elim:   bool = True
    premios_apostados: Dict[str, str] = field(default_factory=dict)
    premios_aciertos:  List[str]      = field(default_factory=list)

    @property
    def total(self) -> int:
        return (self.puntos_grupos + self.puntos_elim
                + self.puntos_quien_pasa + self.puntos_especiales)


# ─────────────────────── lectura de Excels ─────────────────────────
def _abrir(libro: Path):
    return load_workbook(libro, data_only=True, read_only=True)


def _leer_hoja_apuestas(ws, es_eliminatoria: bool) -> tuple:
    """
    Devuelve (apuestas, quien_pasa).
    - apuestas:   id_partido -> (gl, gv)
    - quien_pasa: id_partido -> equipo (solo si es_eliminatoria y columna G rellena)
    """
    apuestas: Dict[str, tuple] = {}
    quien_pasa: Dict[str, str] = {}

    for fila in ws.iter_rows(min_row=1, values_only=True):
        f = list(fila) + [None] * (10 - len(fila))
        id_part = f[2]
        gl, gv  = f[3], f[4]
        if not isinstance(id_part, str):
            continue
        id_part = id_part.strip()
        if not (id_part.startswith("G-") or id_part.startswith("R32-")
                or id_part.startswith("R16-") or id_part.startswith("QF-")
                or id_part in ("SF-01", "SF-02", "TER", "FIN")):
            continue
        if isinstance(gl, (int, float)) and isinstance(gv, (int, float)):
            apuestas[id_part] = (int(gl), int(gv))
        # Quién pasa: columna G (índice 6) solo en eliminatorias
        if es_eliminatoria:
            qp = f[6]
            if isinstance(qp, str) and qp.strip():
                quien_pasa[id_part] = qp.strip()

    return apuestas, quien_pasa


def _leer_alias(ws) -> Optional[str]:
    for ref in ("C2", "B2"):
        v = ws[ref].value
        if isinstance(v, str) and v.strip() and v.strip().lower() not in {
                "alias", "alias / nombre de apuesta:"}:
            return v.strip()
    return None


def _leer_premios_especiales(wb) -> Dict[str, str]:
    """Lee la hoja 'Premios' del libro: MVP en C3, Goleador en C4, Portero en C5."""
    if "Premios" not in wb.sheetnames:
        return {}
    ws = wb["Premios"]
    premios = {}
    mapa = {3: "mvp", 4: "goleador", 5: "portero"}
    for fila, clave in mapa.items():
        v = ws.cell(row=fila, column=3).value
        if isinstance(v, str) and v.strip():
            premios[clave] = v.strip()
    return premios


_HOJAS_VALIDAS = {"Fase Grupos", "Eliminatorias"}


def leer_pronostico(libro: Path, limites: dict) -> Optional[Pronostico]:
    try:
        wb = _abrir(libro)
    except Exception as e:
        print(f"⚠ No se pudo leer {libro.name}: {e}", file=sys.stderr)
        return None

    alias = None
    apuestas: Dict[str, tuple] = {}
    quien_pasa: Dict[str, str] = {}

    for hoja in wb.sheetnames:
        if hoja not in _HOJAS_VALIDAS:
            continue
        ws = wb[hoja]
        if not alias:
            alias = _leer_alias(ws)
        es_elim = (hoja == "Eliminatorias")
        ap, qp = _leer_hoja_apuestas(ws, es_eliminatoria=es_elim)
        apuestas.update(ap)
        quien_pasa.update(qp)

    premios_especiales = _leer_premios_especiales(wb)
    wb.close()

    if not alias:
        print(f"⚠ {libro.name} no tiene alias en B2/C2 — se ignora")
        return None

    mtime = datetime.fromtimestamp(libro.stat().st_mtime, tz=timezone.utc)
    p = Pronostico(
        alias=alias, archivo=libro.name, modified_time=mtime,
        apuestas=apuestas, quien_pasa=quien_pasa,
        premios_especiales=premios_especiales,
    )

    try:
        lim_g = datetime.fromisoformat(limites["limite_grupos"])
        lim_e = datetime.fromisoformat(limites.get("limite_eliminatorias",
                                       limites.get("limite_ronda_32", "2099-01-01T00:00:00+00:00")))
        if mtime > lim_g:
            p.valido_grupos = False
        if mtime > lim_e:
            p.valido_elim = False
    except Exception:
        pass

    return p


def _leer_resultados_de_archivo(archivo: Path) -> Dict[str, Resultado]:
    resultados: Dict[str, Resultado] = {}
    if not archivo.exists():
        return resultados
    wb = _abrir(archivo)
    for hoja in wb.sheetnames:
        if hoja not in _HOJAS_VALIDAS:
            continue
        ws = wb[hoja]
        for fila in ws.iter_rows(min_row=1, values_only=True):
            f = list(fila) + [None] * (10 - len(fila))
            id_part, gl, gv = f[2], f[3], f[4]
            if not isinstance(id_part, str):
                continue
            id_part = id_part.strip()
            if isinstance(gl, (int, float)) and isinstance(gv, (int, float)):
                resultados[id_part] = Resultado(real=(int(gl), int(gv)))
    wb.close()
    return resultados


def leer_resultados_reales() -> Dict[str, Resultado]:
    """Fusiona resultados de grupos (resultados_reales.xlsx) y eliminatorias (resultados_reales_elim.xlsx)."""
    r = _leer_resultados_de_archivo(ARCHIVO_REALES)
    r.update(_leer_resultados_de_archivo(ARCHIVO_REALES_ELIM))
    return r


def leer_premios_reales() -> Dict[str, str]:
    """Lee data/premios_reales.json → {"mvp": "...", "goleador": "...", "portero": "..."}"""
    if not ARCHIVO_PREMIOS_R.exists():
        return {}
    try:
        return json.loads(ARCHIVO_PREMIOS_R.read_text(encoding="utf-8"))
    except Exception:
        return {}


def leer_quien_paso() -> Dict[str, str]:
    """
    Lee data/quien_paso.json → {"R32-01": "España", ...}
    Solo es necesario para partidos que acabaron en empate a 90' (prórroga/penaltis).
    Para partidos no empatados, el ganador se deduce del marcador.
    """
    if not ARCHIVO_QUIEN_P.exists():
        return {}
    try:
        return json.loads(ARCHIVO_QUIEN_P.read_text(encoding="utf-8"))
    except Exception:
        return {}


# ─────────────────────── puntuación ────────────────────────────────
def signo(a: int, b: int) -> int:
    if a > b:  return  1
    if a < b:  return -1
    return 0


def evaluar(apostado: Optional[tuple], real: Optional[tuple]):
    if real is None:
        return 0, "pendiente"
    if apostado is None:
        return 0, "no_apostado"
    if apostado == real:
        return PUNTOS_PLENO, "pleno"
    if signo(*apostado) == signo(*real):
        return PUNTOS_TENDENCIA, "tendencia"
    return 0, "fallo"


def quien_pasa_real_para(id_partido: str,
                          resultado: Optional[tuple],
                          partidos_meta: dict,
                          quien_paso: Dict[str, str]) -> Optional[str]:
    """
    Devuelve el nombre del equipo que efectivamente pasó, o None si aún no se sabe.
    - Ganador claro: se deduce del marcador.
    - Empate a 90': se busca en quien_paso.json.
    """
    if resultado is None:
        return None
    gl, gv = resultado
    meta = partidos_meta.get(id_partido, {})
    if gl > gv:
        return meta.get("local")
    if gv > gl:
        return meta.get("visitante")
    return quien_paso.get(id_partido)


def calcular(participantes: Dict[str, Participante],
             partidos: List[dict],
             resultados: Dict[str, Resultado],
             premios_reales: Dict[str, str],
             quien_paso: Dict[str, str]):
    """Rellena puntos y detalle de cada participante."""

    meta = {p["id"]: p for p in partidos}

    for part in participantes.values():
        for p in partidos:
            id_part = p["id"]
            fase    = p["fase"]
            real    = resultados.get(id_part, Resultado()).real
            es_elim = fase in FASES_ELIMINATORIAS

            if not es_elim and not part.valido_grupos:
                apostado    = None
                qp_apostado = None
            elif es_elim and not part.valido_elim:
                apostado    = None
                qp_apostado = None
            else:
                apostado    = part._apuestas.get(id_part)
                qp_apostado = part._quien_pasa.get(id_part) if es_elim else None

            puntos, tipo = evaluar(apostado, real)

            pts_qp  = 0
            qp_real = None
            if es_elim and real is not None:
                qp_real = quien_pasa_real_para(id_part, real, meta, quien_paso)
                if qp_real and qp_apostado:
                    if qp_apostado.strip().lower() == qp_real.strip().lower():
                        pts_qp = PUNTOS_QUIEN_PASA

            part.detalle.append(DetalleAcierto(
                id_part, apostado, real, puntos, tipo,
                quien_pasa_apostado=qp_apostado,
                quien_pasa_real=qp_real,
                puntos_quien_pasa=pts_qp,
            ))

            if not es_elim:
                part.puntos_grupos += puntos
            else:
                part.puntos_elim       += puntos
                part.puntos_quien_pasa += pts_qp

            if tipo == "pleno":     part.aciertos_pleno      += 1
            if tipo == "tendencia": part.aciertos_tendencia  += 1
            if pts_qp > 0:          part.aciertos_quien_pasa += 1

        # Premios especiales
        for clave in ("mvp", "goleador", "portero"):
            real_val    = premios_reales.get(clave, "").strip().lower()
            apuesta_val = part.premios_apostados.get(clave, "").strip().lower()
            if real_val and apuesta_val and apuesta_val == real_val:
                part.puntos_especiales += PUNTOS_BONUS_ESPECIAL
                part.premios_aciertos.append(clave)


# ─────────────────────── orquestación ──────────────────────────────
def cargar_limites() -> dict:
    if ARCHIVO_LIMITES.exists():
        return json.loads(ARCHIVO_LIMITES.read_text(encoding="utf-8"))
    return {
        "limite_grupos":        "2026-06-11T15:00:00+00:00",
        "limite_eliminatorias": "2026-06-28T15:00:00+00:00",
    }


def main():
    if not ARCHIVO_PARTIDOS.exists():
        print("✘ Falta data/matches.json — ejecuta antes scripts/build_matches.py", file=sys.stderr)
        sys.exit(1)

    datos    = json.loads(ARCHIVO_PARTIDOS.read_text(encoding="utf-8"))
    partidos = datos["partidos"]
    limites  = cargar_limites()

    avatares: Dict[str, str] = {}
    if ARCHIVO_AVATARES.exists():
        avatares = json.loads(ARCHIVO_AVATARES.read_text(encoding="utf-8"))

    banderas = {}
    if ARCHIVO_BANDERAS.exists():
        banderas = json.loads(ARCHIVO_BANDERAS.read_text(encoding="utf-8"))
    for p in partidos:
        p["bandera_local"]     = banderas.get(p.get("local", ""), "")
        p["bandera_visitante"] = banderas.get(p.get("visitante", ""), "")

    # 1. Leer pronósticos
    #    Convención de nombres en pronosticos/:
    #      porra_ALIAS.xlsx      → datos de Fase de Grupos + Premios
    #      porra_ALIAS_elim.xlsx → datos de Eliminatorias
    #
    #    También se acepta cualquier *.xlsx que no empiece por ~$ (compatibilidad).
    #    Si hay varios archivos para el mismo alias, se queda el más reciente por tipo.

    DIR_PRONO.mkdir(exist_ok=True)

    # Separar archivos de grupos y de eliminatorias
    grupos_por_alias: Dict[str, Pronostico]  = {}
    elim_por_alias:   Dict[str, Pronostico]  = {}

    for libro in sorted(DIR_PRONO.glob("*.xlsx")):
        if libro.name.startswith("~$"):
            continue
        es_elim_file = libro.stem.endswith("_elim")
        pr = leer_pronostico(libro, limites)
        if not pr:
            continue
        if es_elim_file:
            bucket = elim_por_alias
        else:
            bucket = grupos_por_alias
        if pr.alias not in bucket or pr.modified_time > bucket[pr.alias].modified_time:
            bucket[pr.alias] = pr

    # Todos los alias conocidos (pueden tener solo grupos, solo elim, o ambos)
    todos_alias = set(grupos_por_alias) | set(elim_por_alias)

    if not todos_alias:
        print(f"ℹ Sin pronósticos en {DIR_PRONO} todavía — se generará data.json vacío")

    # 2. Construir participantes fusionando ambos archivos
    participantes: Dict[str, Participante] = {}
    for alias in todos_alias:
        pr_g = grupos_por_alias.get(alias)
        pr_e = elim_por_alias.get(alias)

        # El archivo de referencia principal es el de grupos (para mtime y premios)
        pr_ref = pr_g or pr_e

        apuestas   = {}
        quien_pasa = {}
        premios    = {}

        if pr_g:
            apuestas.update(pr_g.apuestas)
            premios = pr_g.premios_especiales
        if pr_e:
            apuestas.update(pr_e.apuestas)
            quien_pasa.update(pr_e.quien_pasa)

        p = Participante(
            alias=alias,
            archivo=pr_ref.archivo,
            modified_time=pr_ref.modified_time,
            valido_grupos=pr_g.valido_grupos if pr_g else False,
            valido_elim=pr_e.valido_elim if pr_e else False,
            premios_apostados=premios,
        )
        p._apuestas   = apuestas
        p._quien_pasa = quien_pasa
        participantes[alias] = p

    # 3. Resultados, premios y quién pasó
    resultados     = leer_resultados_reales()
    premios_reales = leer_premios_reales()
    quien_paso     = leer_quien_paso()

    calcular(participantes, partidos, resultados, premios_reales, quien_paso)

    # 4. Serializar para el dashboard
    ranking = sorted(
        participantes.values(),
        key=lambda x: (x.total, x.aciertos_pleno),
        reverse=True,
    )

    # ── Auto-asignación de avatares ──────────────────────────────────
    import random as _random
    _AVATARES_DISPONIBLES = sorted([
        f for f in (RAIZ / "Avatar").iterdir()
        if f.suffix.lower() == ".png"
    ], key=lambda f: f.name)
    if _AVATARES_DISPONIBLES:
        _usados = set(avatares.values())
        _libres = [f.name for f in _AVATARES_DISPONIBLES if f.name not in _usados]
        _todos  = [f.name for f in _AVATARES_DISPONIBLES]
        _modificado = False
        for p in ranking:
            if p.alias not in avatares:
                if _libres:
                    elegido = _random.choice(_libres)
                    _libres.remove(elegido)
                else:
                    # Todos asignados: reutilizar aleatoriamente
                    elegido = _random.choice(_todos)
                avatares[p.alias] = elegido
                _modificado = True
        if _modificado:
            ARCHIVO_AVATARES.write_text(
                json.dumps(avatares, ensure_ascii=False, indent=2), encoding="utf-8"
            )

    # ── Precargar avatares como base64 para el standalone ───────────
    import base64 as _base64
    _DIR_AVATAR = RAIZ / "Avatar"
    _avatar_b64_cache: Dict[str, str] = {}
    for _alias, _fname in avatares.items():
        _fpath = _DIR_AVATAR / _fname
        if _fpath.exists():
            _data = _base64.b64encode(_fpath.read_bytes()).decode("ascii")
            _avatar_b64_cache[_alias] = f"data:image/png;base64,{_data}"

    salida = {
        "torneo": datos.get("torneo", "Mundial 2026"),
        "actualizado": datetime.now(timezone.utc).isoformat(timespec="seconds"),
        "limites": limites,
        "puntuacion": {
            "pleno":          PUNTOS_PLENO,
            "tendencia":      PUNTOS_TENDENCIA,
            "quien_pasa":     PUNTOS_QUIEN_PASA,
            "bonus_especial": PUNTOS_BONUS_ESPECIAL,
        },
        "partidos": partidos,
        "resultados": {k: list(v.real) if v.real else None for k, v in resultados.items()},
        "premios_reales": premios_reales,
        "ranking": [
            {
                "alias":             p.alias,
                "avatar":            avatares.get(p.alias, ""),
                "avatar_b64":        _avatar_b64_cache.get(p.alias, ""),
                "archivo":           p.archivo,
                "modified":          p.modified_time.isoformat(timespec="seconds"),
                "puntos_grupos":     p.puntos_grupos,
                "puntos_elim":       p.puntos_elim,
                "puntos_quien_pasa": p.puntos_quien_pasa,
                "puntos_especiales": p.puntos_especiales,
                "total":             p.total,
                "pleno":             p.aciertos_pleno,
                "tendencia":         p.aciertos_tendencia,
                "quien_pasa":        p.aciertos_quien_pasa,
                "valido_grupos":     p.valido_grupos,
                "valido_elim":       p.valido_elim,
                "premios_apostados": p.premios_apostados,
                "premios_aciertos":  p.premios_aciertos,
                "detalle": [
                    {
                        "id":              d.id_partido,
                        "apostado":        list(d.apostado) if d.apostado else None,
                        "real":            list(d.real) if d.real else None,
                        "puntos":          d.puntos,
                        "tipo":            d.tipo,
                        "quien_pasa_ap":   d.quien_pasa_apostado,
                        "quien_pasa_real": d.quien_pasa_real,
                        "pts_quien_pasa":  d.puntos_quien_pasa,
                    } for d in p.detalle
                ],
            } for p in ranking
        ],
    }

    DIR_DASHBOARD.mkdir(exist_ok=True)
    ARCHIVO_SALIDA.write_text(json.dumps(salida, ensure_ascii=False, indent=2), encoding="utf-8")
    print(f"OK · {len(participantes)} participantes · "
          f"{sum(1 for r in resultados.values() if r.real)} resultados procesados")
    print(f"    → {ARCHIVO_SALIDA}")

    # ── Dashboard (web + standalone) ───────────────────────────
    _base = DIR_DASHBOARD / "index.html"
    if _base.exists():
        import re as _re
        _src = _base.read_text(encoding="utf-8")
        # Elimina inyección previa
        _src = _re.sub(
            r"<!-- DATOS EMBEBIDOS POR scorer\.py · NO EDITAR A MANO -->\n<script>window\.PRECARGA = .+?;</script>\n",
            "", _src, flags=_re.DOTALL
        )
        precarga = json.dumps(salida, ensure_ascii=False).replace("</", "<\/")
        inyeccion = (
            "<!-- DATOS EMBEBIDOS POR scorer.py · NO EDITAR A MANO -->\n"
            f"<script>window.PRECARGA = {precarga};</script>\n"
        )
        html = _src.replace("</head>", inyeccion + "</head>", 1) if "</head>" in _src else inyeccion + _src
        _base.write_text(html, encoding="utf-8")
        (DIR_DASHBOARD / "dashboard.html").write_text(html, encoding="utf-8")
        print(f"    → {_base}  (web + local)")


if __name__ == "__main__":
    main()
