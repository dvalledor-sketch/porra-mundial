"""
Carga datos de DEMO para ver el dashboard funcionando:
- 4 pronósticos ficticios (Daniel, Lola, Pablo, Marta) en pronosticos/
- 5 resultados reales del Grupo A en data/resultados_reales.xlsx
Después ejecuta scorer.py automáticamente.

Para volver al estado de producción (en blanco), ejecuta:
    python3 scripts/limpiar_demo.py
"""
from __future__ import annotations
import shutil
import subprocess
import sys
from pathlib import Path
from openpyxl import load_workbook

RAIZ = Path(__file__).resolve().parent.parent
DIR_PLANTILLAS = RAIZ / "plantillas"
DIR_PRONO      = RAIZ / "pronosticos"
DIR_DATA       = RAIZ / "data"


def rellenar(alias: str, archivo: Path, apuestas: dict):
    wb = load_workbook(archivo)
    # Alias en las hojas que lo usan (C2 es el top-left de la celda fusionada)
    # Premios usa B2 como celda principal, pero el scorer sólo lee de Fase Grupos
    for hoja in ("Fase Grupos", "Eliminatorias"):
        if hoja in wb.sheetnames:
            wb[hoja]["C2"] = alias
    # Rellenar goles en Fase Grupos
    if "Fase Grupos" in wb.sheetnames:
        ws = wb["Fase Grupos"]
        for fila in ws.iter_rows(min_row=1):
            id_part = fila[2].value
            if isinstance(id_part, str) and id_part in apuestas:
                gl, gv = apuestas[id_part]
                fila[3].value = gl
                fila[4].value = gv
    wb.save(archivo)


def rellenar_reales(reales: dict):
    archivo = DIR_DATA / "resultados_reales.xlsx"
    wb = load_workbook(archivo)
    ws = wb["Fase Grupos"]
    for fila in ws.iter_rows(min_row=1):
        id_part = fila[2].value
        if isinstance(id_part, str) and id_part in reales:
            gl, gv = reales[id_part]
            fila[3].value = gl
            fila[4].value = gv
    wb.save(archivo)


def main():
    REALES = {
        "G-A1-A2": (2, 1),  # México 2-1 Sudáfrica
        "G-A3-A4": (0, 0),  # Corea del Sur 0-0 Chequia
        "G-A1-A3": (3, 0),  # México 3-0 Corea del Sur
        "G-A2-A4": (1, 2),  # Sudáfrica 1-2 Chequia
        "G-A1-A4": (1, 1),  # México 1-1 Chequia
    }
    PARTICIPANTES = [
        ("Daniel", {"G-A1-A2": (2,1), "G-A3-A4": (0,0), "G-A1-A3": (3,0), "G-A2-A4": (1,2), "G-A1-A4": (2,1)}),
        ("Lola",   {"G-A1-A2": (0,3), "G-A3-A4": (2,1), "G-A1-A3": (0,1), "G-A2-A4": (0,5), "G-A1-A4": (1,1)}),
        ("Pablo",  {"G-A1-A2": (4,2), "G-A3-A4": (1,1), "G-A1-A3": (2,0), "G-A2-A4": (0,0), "G-A1-A4": (3,3)}),
        ("Marta",  {"G-A1-A2": (1,1), "G-A3-A4": (2,2), "G-A1-A3": (1,0), "G-A2-A4": (1,3), "G-A1-A4": (0,2)}),
    ]

    DIR_PRONO.mkdir(exist_ok=True)
    rellenar_reales(REALES)

    for alias, apuestas in PARTICIPANTES:
        archivo = DIR_PRONO / f"demo_{alias.lower()}.xlsx"
        shutil.copy(DIR_PLANTILLAS / "plantilla_porra.xlsx", archivo)
        rellenar(alias, archivo, apuestas)
        print(f"  ✓ {archivo.name}")

    print()
    subprocess.run([sys.executable, str(RAIZ / "scripts" / "scorer.py")], cwd=RAIZ, check=True)
    print()
    print("✔ DEMO cargada. Abre dashboard/index.html para verlo.")


if __name__ == "__main__":
    main()
